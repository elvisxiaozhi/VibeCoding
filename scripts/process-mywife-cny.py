#!/usr/bin/env python3

from __future__ import annotations

import csv
import json
import warnings
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path

import openpyxl


warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    module="openpyxl.styles.stylesheet",
)


ROOT_DIR = Path.cwd()
INPUT_DIR = ROOT_DIR / "deploy" / "my wife"
OUTPUT_DIR = ROOT_DIR / "deploy" / "mywife_processed"

REPORT_PATH = OUTPUT_DIR / "report.md"
CSV_PATH = OUTPUT_DIR / "cashflows.csv"
JSON_PATH = OUTPUT_DIR / "parsed.json"


@dataclass
class Flow:
    account: str
    flow_type: str
    record_type: str
    date: str
    amount: float
    note: str
    detail: str
    file_name: str


@dataclass
class Snapshot:
    account: str
    date: str
    total_asset: float
    file_name: str


@dataclass
class AccountMeta:
    account: str
    target: str
    expected_return: str
    investment_horizon: str
    currency: str
    bucket: str
    file_name: str


def normalize_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if value is None:
        return ""
    return str(value)


def money(value: float) -> str:
    return f"{value:.2f}"


def pct(value: float) -> str:
    return f"{value * 100:.2f}%"


def detect_flow_type(amount: float, note: str) -> str:
    if amount > 0:
        return "contribution"
    if "分红" in note:
        return "dividend"
    return "withdrawal"


def load_workbook_data(path: Path) -> tuple[AccountMeta, list[Flow], list[Snapshot]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    meta = AccountMeta(
        account=str(rows[1][0] or ""),
        target=str(rows[1][1] or ""),
        expected_return=str(rows[1][2] or ""),
        investment_horizon=str(rows[1][3] or ""),
        currency=str(rows[1][4] or ""),
        bucket=str(rows[1][5] or ""),
        file_name=path.name,
    )

    flows: list[Flow] = []
    snapshots: list[Snapshot] = []

    for row in rows[4:]:
        if not any(cell not in (None, "") for cell in row):
            continue

        record_type = str(row[0] or "")
        date = normalize_date(row[1])
        amount = row[2]
        total_asset = row[3]
        note = str(row[4] or "")
        detail = str(row[6] or "")

        if record_type == "转入转出" and amount is not None:
            flow_amount = float(amount)
            flows.append(
                Flow(
                    account=meta.account,
                    flow_type=detect_flow_type(flow_amount, note),
                    record_type=record_type,
                    date=date,
                    amount=abs(flow_amount),
                    note=note,
                    detail=detail,
                    file_name=path.name,
                )
            )
        elif record_type == "记总资产" and total_asset is not None:
            snapshots.append(
                Snapshot(
                    account=meta.account,
                    date=date,
                    total_asset=float(total_asset),
                    file_name=path.name,
                )
            )

    return meta, flows, snapshots


def latest_snapshot(snapshots: list[Snapshot]) -> Snapshot | None:
    if not snapshots:
        return None
    return max(snapshots, key=lambda item: item.date)


def main() -> None:
    if not INPUT_DIR.exists():
        raise SystemExit(f"Missing input directory: {INPUT_DIR}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    workbook_paths = sorted(INPUT_DIR.glob("*.xlsx"))
    if not workbook_paths:
        raise SystemExit(f"No xlsx files found in {INPUT_DIR}")

    metas: list[AccountMeta] = []
    flows: list[Flow] = []
    snapshots: list[Snapshot] = []

    for workbook_path in workbook_paths:
        meta, workbook_flows, workbook_snapshots = load_workbook_data(workbook_path)
        metas.append(meta)
        flows.extend(workbook_flows)
        snapshots.extend(workbook_snapshots)

    flows.sort(key=lambda item: (item.date, item.account, item.flow_type, item.file_name))
    snapshots.sort(key=lambda item: (item.date, item.account, item.file_name))

    meta_by_account = {meta.account: meta for meta in metas}
    account_summaries = []

    total_contribution = 0.0
    total_withdrawal = 0.0
    total_dividend = 0.0
    total_latest_asset = 0.0

    for account in sorted(meta_by_account):
        account_flows = [flow for flow in flows if flow.account == account]
        account_snapshots = [snapshot for snapshot in snapshots if snapshot.account == account]
        latest = latest_snapshot(account_snapshots)

        contribution = sum(flow.amount for flow in account_flows if flow.flow_type == "contribution")
        withdrawal = sum(flow.amount for flow in account_flows if flow.flow_type == "withdrawal")
        dividend = sum(flow.amount for flow in account_flows if flow.flow_type == "dividend")
        latest_asset = latest.total_asset if latest else 0.0
        total_cash_return = withdrawal + dividend
        net_invested = contribution - total_cash_return
        total_pnl = latest_asset + total_cash_return - contribution
        total_return_rate = (total_pnl / contribution) if contribution else 0.0

        total_contribution += contribution
        total_withdrawal += withdrawal
        total_dividend += dividend
        total_latest_asset += latest_asset

        account_summaries.append(
            {
                "account": account,
                "expected_return": meta_by_account[account].expected_return,
                "investment_horizon": meta_by_account[account].investment_horizon,
                "bucket": meta_by_account[account].bucket,
                "currency": meta_by_account[account].currency,
                "contribution": round(contribution, 2),
                "withdrawal": round(withdrawal, 2),
                "dividend": round(dividend, 2),
                "cash_return": round(total_cash_return, 2),
                "net_invested": round(net_invested, 2),
                "latest_asset": round(latest_asset, 2),
                "latest_asset_date": latest.date if latest else "",
                "snapshot_count": len(account_snapshots),
                "total_pnl": round(total_pnl, 2),
                "total_return_rate": total_return_rate,
                "file_name": meta_by_account[account].file_name,
            }
        )

    total_cash_return = total_withdrawal + total_dividend
    total_net_invested = total_contribution - total_cash_return
    total_pnl = total_latest_asset + total_cash_return - total_contribution
    total_return_rate = (total_pnl / total_contribution) if total_contribution else 0.0
    latest_snapshot_date = max((snapshot.date for snapshot in snapshots), default="")

    report_lines = [
        "# My Wife 人民币资产整理报表",
        "",
        "## 说明",
        "",
        "- 数据来源：`deploy/my wife/*.xlsx`。",
        "- 资产币种：人民币（CNY）。",
        "- 口径：正数 `转入转出金额` 记为转入，负数且 `投资日志` 含 `分红` 记为股息，其他负数记为普通转出/赎回。",
        "- 总收益口径：`最新总资产 + 全部已回笼现金 - 累计转入`。",
        "",
        "## 汇总",
        "",
        "| 指标 | 数值 |",
        "| --- | --- |",
        f"| Excel 文件数 | {len(workbook_paths)} |",
        f"| 账户数 | {len(account_summaries)} |",
        f"| 累计转入 | {money(total_contribution)} CNY |",
        f"| 普通转出/赎回 | {money(total_withdrawal)} CNY |",
        f"| 已确认股息 | {money(total_dividend)} CNY |",
        f"| 已回笼现金合计 | {money(total_cash_return)} CNY |",
        f"| 当前最新总资产 | {money(total_latest_asset)} CNY |",
        f"| 当前净投入 | {money(total_net_invested)} CNY |",
        f"| 累计总收益 | {money(total_pnl)} CNY |",
        f"| 累计总收益率 | {pct(total_return_rate)} |",
        f"| 最新资产记录日期 | {latest_snapshot_date or '未知'} |",
        "",
        "## 账户汇总",
        "",
        "| 账户 | 最新总资产日期 | 累计转入 | 普通转出/赎回 | 已确认股息 | 当前净投入 | 最新总资产 | 累计总收益 | 累计总收益率 |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- |",
        *[
            f"| {item['account']} | {item['latest_asset_date']} | {money(item['contribution'])} | {money(item['withdrawal'])} | {money(item['dividend'])} | {money(item['net_invested'])} | {money(item['latest_asset'])} | {money(item['total_pnl'])} | {pct(item['total_return_rate'])} |"
            for item in account_summaries
        ],
        "",
        "## 现金流明细",
        "",
        "| 日期 | 账户 | 类型 | 金额 | 备注 | 来源文件 |",
        "| --- | --- | --- | --- | --- | --- |",
        *[
            f"| {flow.date} | {flow.account} | {flow.flow_type} | {money(flow.amount)} | {flow.note or ''} | {flow.file_name} |"
            for flow in flows
        ],
        "",
        "## 结论",
        "",
        f"- 这批人民币资产当前最新总资产合计 {money(total_latest_asset)} CNY，累计已回笼现金 {money(total_cash_return)} CNY，累计总收益 {money(total_pnl)} CNY。",
        f"- 当前净投入为 {money(total_net_invested)} CNY；按当前口径，整体累计总收益率为 {pct(total_return_rate)}。",
        f"- 其中最新资产记录最晚的是 {latest_snapshot_date or '未知'}，如果后续 Excel 有更新，重新运行脚本即可刷新结果。",
        "",
    ]

    with CSV_PATH.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["date", "account", "flow_type", "record_type", "amount", "note", "detail", "file_name"])
        for flow in flows:
            writer.writerow([
                flow.date,
                flow.account,
                flow.flow_type,
                flow.record_type,
                f"{flow.amount:.2f}",
                flow.note,
                flow.detail,
                flow.file_name,
            ])

    parsed = {
        "source_dir": str(INPUT_DIR.relative_to(ROOT_DIR)),
        "output_dir": str(OUTPUT_DIR.relative_to(ROOT_DIR)),
        "summary": {
            "file_count": len(workbook_paths),
            "account_count": len(account_summaries),
            "total_contribution": round(total_contribution, 2),
            "total_withdrawal": round(total_withdrawal, 2),
            "total_dividend": round(total_dividend, 2),
            "total_cash_return": round(total_cash_return, 2),
            "total_latest_asset": round(total_latest_asset, 2),
            "total_net_invested": round(total_net_invested, 2),
            "total_pnl": round(total_pnl, 2),
            "total_return_rate": total_return_rate,
            "latest_snapshot_date": latest_snapshot_date,
        },
        "accounts": account_summaries,
        "meta": [asdict(meta) for meta in metas],
        "flows": [asdict(flow) for flow in flows],
        "snapshots": [asdict(snapshot) for snapshot in snapshots],
    }

    REPORT_PATH.write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    JSON_PATH.write_text(json.dumps(parsed, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(
        f"Generated {REPORT_PATH.relative_to(ROOT_DIR)}, "
        f"{CSV_PATH.relative_to(ROOT_DIR)}, and {JSON_PATH.relative_to(ROOT_DIR)}"
    )


if __name__ == "__main__":
    main()
